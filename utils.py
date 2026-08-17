"""업로드된 중학교 교육과정 편제표(XLSM)를 시트별 점검용 구조로 변환한다."""

from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl

TARGET_SHEETS = ("1학년", "2학년", "3학년", "학년전체")
GRADE_SHEETS = ("1학년", "2학년", "3학년")

# 각 시트는 동일한 6개 학기를 표현하지만, 셀 배치가 다르다.
SEMESTER_COLUMNS: dict[str, list[dict[str, str]]] = {
    "1학년": [
        {"label": "1학년 1학기", "planned": "M", "free_adjust": "N", "autonomy_adjust": "O", "operation": "P"},
        {"label": "1학년 2학기", "planned": "Q", "free_adjust": "R", "autonomy_adjust": "S", "operation": "T"},
        {"label": "2학년 1학기", "planned": "U", "autonomy_adjust": "V", "operation": "W"},
        {"label": "2학년 2학기", "planned": "X", "autonomy_adjust": "Y", "operation": "Z"},
        {"label": "3학년 1학기", "planned": "AA", "autonomy_adjust": "AB", "operation": "AC"},
        {"label": "3학년 2학기", "planned": "AD", "autonomy_adjust": "AE", "operation": "AF"},
    ],
    "2학년": [
        {"label": "1학년 1학기", "planned": "M", "free_adjust": "N", "autonomy_adjust": "O", "operation": "P"},
        {"label": "1학년 2학기", "planned": "Q", "free_adjust": "R", "autonomy_adjust": "S", "operation": "T"},
        {"label": "2학년 1학기", "planned": "U", "autonomy_adjust": "V", "operation": "W"},
        {"label": "2학년 2학기", "planned": "X", "autonomy_adjust": "Y", "operation": "Z"},
        {"label": "3학년 1학기", "planned": "AA", "autonomy_adjust": "AB", "operation": "AC"},
        {"label": "3학년 2학기", "planned": "AD", "autonomy_adjust": "AE", "operation": "AF"},
    ],
    "3학년": [
        {"label": "1학년 1학기", "planned": "M", "free_adjust": "N", "autonomy_adjust": "O", "operation": "P"},
        {"label": "1학년 2학기", "planned": "Q", "free_adjust": "R", "autonomy_adjust": "S", "operation": "T"},
        {"label": "2학년 1학기", "planned": "U", "autonomy_adjust": "V", "operation": "W"},
        {"label": "2학년 2학기", "planned": "X", "autonomy_adjust": "Y", "operation": "Z"},
        {"label": "3학년 1학기", "planned": "AA", "autonomy_adjust": "AB", "operation": "AC"},
        {"label": "3학년 2학기", "planned": "AD", "autonomy_adjust": "AE", "operation": "AF"},
    ],
    "학년전체": [
        {"label": "1학년 1학기", "planned": "I", "free_adjust": "J", "autonomy_adjust": "K", "operation": "L"},
        {"label": "1학년 2학기", "planned": "M", "free_adjust": "N", "autonomy_adjust": "O", "operation": "P"},
        {"label": "2학년 1학기", "planned": "Q", "autonomy_adjust": "R", "operation": "S"},
        {"label": "2학년 2학기", "planned": "T", "autonomy_adjust": "U", "operation": "V"},
        {"label": "3학년 1학기", "planned": "W", "autonomy_adjust": "X", "operation": "Y"},
        {"label": "3학년 2학기", "planned": "Z", "autonomy_adjust": "AA", "operation": "AB"},
    ],
}


def clean_text(value: Any) -> str:
    """셀 문자열을 비교·표시에 적합하도록 정리한다."""
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def as_number(value: Any) -> float:
    """숫자 셀을 0 기본값의 float으로 변환한다."""
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def display_number(value: float) -> int | float:
    """정수처럼 보이는 값은 int로 반환한다."""
    return int(value) if float(value).is_integer() else round(float(value), 2)


def source_to_bytes(source: Any) -> bytes:
    """Streamlit UploadedFile 또는 파일 경로를 openpyxl 입력 바이트로 변환한다."""
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    if hasattr(source, "getvalue"):
        return source.getvalue()
    if hasattr(source, "read"):
        try:
            source.seek(0)
        except (AttributeError, OSError):
            pass
        return source.read()
    raise TypeError("지원하지 않는 파일 입력입니다. 경로 또는 업로드 파일을 전달하세요.")


def values_by_semester(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, sheet_name: str) -> dict[str, dict[str, int | float]]:
    """한 행의 학기별 편성·조정·운영 값을 공통 구조로 읽는다."""
    values: dict[str, dict[str, int | float]] = {}
    for config in SEMESTER_COLUMNS[sheet_name]:
        item: dict[str, int | float] = {}
        for field in ("planned", "free_adjust", "autonomy_adjust", "operation"):
            column = config.get(field)
            if column:
                item[field] = display_number(as_number(ws[f"{column}{row}"].value))
            else:
                item[field] = 0
        values[config["label"]] = item
    return values


def operation_total(semesters: dict[str, dict[str, int | float]]) -> int | float:
    return display_number(sum(as_number(item.get("operation")) for item in semesters.values()))


def operation_by_semester(semesters: dict[str, dict[str, int | float]]) -> dict[str, int | float]:
    return {semester: display_number(as_number(values.get("operation"))) for semester, values in semesters.items()}


def find_collaborative_art_hours(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | float:
    """협력종합예술활동 표에서 운영 시수를 찾아 반환한다.

    학년 시트와 학년전체 시트는 안내 문구와 실제 '운영 시수' 행이 서로 다른
    열에 배치되어 있다. 따라서 문구를 찾은 뒤 인접 행의 '운영 시수' 라벨을
    찾고, 해당 행 전체에서 수치값을 읽는다.
    """
    for row in ws.iter_rows():
        for cell in row:
            if "협력종합예술" not in clean_text(cell.value):
                continue
            for search_row in range(cell.row, min(ws.max_row, cell.row + 8) + 1):
                labels = [clean_text(ws.cell(search_row, col).value) for col in range(1, ws.max_column + 1)]
                if not any("운영 시수" in label for label in labels):
                    continue
                candidates = [
                    float(ws.cell(search_row, col).value)
                    for col in range(1, ws.max_column + 1)
                    if isinstance(ws.cell(search_row, col).value, (int, float))
                    and 1 <= float(ws.cell(search_row, col).value) <= 500
                ]
                if candidates:
                    return display_number(max(candidates))
    return 0


def extract_rows(ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str, start_row: int, end_row: int, include_subjects: bool = False) -> list[dict[str, Any]]:
    """교과·창체·자율시간 행을 단일 레코드 형식으로 추출한다."""
    rows: list[dict[str, Any]] = []
    current_category = ""
    current_group = ""

    for row_number in range(start_row, end_row + 1):
        category = clean_text(ws.cell(row_number, 1).value)
        group = clean_text(ws.cell(row_number, 2).value)
        subject = clean_text(ws.cell(row_number, 3).value)
        detail = clean_text(ws.cell(row_number, 4).value)

        if category:
            current_category = category
        if group:
            current_group = group
        if include_subjects and not subject:
            continue
        if not include_subjects and not (group or subject or detail):
            continue

        semesters = values_by_semester(ws, row_number, sheet_name)
        rows.append(
            {
                "row": row_number,
                "category": current_category,
                "group": current_group,
                "subject": subject,
                "detail": detail,
                "selection_code": clean_text(ws.cell(row_number, 1).value),
                "standard_hours": display_number(as_number(ws.cell(row_number, 6).value)),
                "displayed_planned_hours": display_number(as_number(ws.cell(row_number, 8).value)),
                "displayed_difference_hours": display_number(as_number(ws.cell(row_number, 10).value)),
                "semesters": semesters,
                "operation_hours": operation_total(semesters),
                "operation_by_semester": operation_by_semester(semesters),
            }
        )
    return rows


def extract_sheet_payload(ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> dict[str, Any]:
    """점검에 필요한 특정 행 블록을 시트별 구조화 데이터로 추출한다."""
    subjects = extract_rows(ws, sheet_name, 8, 27, include_subjects=True)
    creative = extract_rows(ws, sheet_name, 29, 32)
    free_semester = extract_rows(ws, sheet_name, 34, 35)
    autonomy_courses = extract_rows(ws, sheet_name, 37, 42)

    course_count = operation_by_semester(values_by_semester(ws, 45, sheet_name))
    autonomy_totals = operation_by_semester(values_by_semester(ws, 43, sheet_name))
    free_semester_totals = operation_by_semester(values_by_semester(ws, 36, sheet_name))

    # 학교스포츠클럽 활동의 순증 여부·방식은 하단 표에 별도 기재된다.
    sports_rows: list[dict[str, Any]] = []
    for row_number in range(55, min(59, ws.max_row) + 1):
        label = clean_text(ws.cell(row_number, 1).value)
        values = []
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row_number, column).value
            if isinstance(value, (int, float)):
                values.append(float(value))
        sports_rows.append({"row": row_number, "label": label, "numeric_values": values})

    return {
        "sheet_name": sheet_name,
        "semesters": [item["label"] for item in SEMESTER_COLUMNS[sheet_name]],
        "subjects": subjects,
        "creative": creative,
        "free_semester": free_semester,
        "autonomy_courses": autonomy_courses,
        "course_count": course_count,
        "autonomy_totals": autonomy_totals,
        "free_semester_totals": free_semester_totals,
        "collaborative_art_hours": find_collaborative_art_hours(ws),
        "sports_rows": sports_rows,
    }


def parse_curriculum_xlsm(source: Any, source_name: str | None = None) -> dict[str, Any]:
    """XLSM 파일 1개를 학교 단위의 시트별 점검 데이터로 변환한다.

    `data_only=True`로 저장된 수식 결과를 읽는다. 따라서 학교는 Excel에서 저장을
    완료한 파일을 제출해야 최신 계산값이 반영된다.
    """
    file_bytes = source_to_bytes(source)
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, keep_vba=True)

    missing_sheets = [sheet for sheet in TARGET_SHEETS if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"필수 시트가 없습니다: {', '.join(missing_sheets)}")

    base_sheet = workbook["기초입력"] if "기초입력" in workbook.sheetnames else None
    metadata = {
        "파일명": source_name or getattr(source, "name", "업로드파일.xlsm"),
        "학년도": base_sheet["C2"].value if base_sheet else None,
        "지원청명": clean_text(base_sheet["C3"].value) if base_sheet else "",
        "설립별": clean_text(base_sheet["C4"].value) if base_sheet else "",
        "학교명": clean_text(base_sheet["C5"].value) if base_sheet else "",
    }
    if not metadata["학교명"]:
        metadata["학교명"] = Path(metadata["파일명"]).stem

    return {
        "metadata": metadata,
        "sheets": {sheet_name: extract_sheet_payload(workbook[sheet_name], sheet_name) for sheet_name in TARGET_SHEETS},
    }


def flatten_curriculum_data(school_payload: dict[str, Any]) -> list[dict[str, Any]]:
    """엑셀 내보내기용 원시 과목 데이터를 행 단위로 펼친다."""
    records: list[dict[str, Any]] = []
    metadata = school_payload["metadata"]
    for sheet_name, payload in school_payload["sheets"].items():
        for subject in payload["subjects"]:
            record = {
                **metadata,
                "점검시트": sheet_name,
                "원본행": subject["row"],
                "대분류": subject["category"],
                "교과군": subject["group"],
                "과목명": subject["subject"],
                "세부내용": subject["detail"],
                "기준시수": subject["standard_hours"],
                "운영시수": subject["operation_hours"],
            }
            record.update({f"{semester}_운영시수": hours for semester, hours in subject["operation_by_semester"].items()})
            records.append(record)
    return records


def clone_payload(payload: dict[str, Any]) -> dict[str, Any]:
    """세션 상태에서 안전하게 재사용할 수 있도록 학교 데이터를 복제한다."""
    return deepcopy(payload)
